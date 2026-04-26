#!/usr/bin/env python3
"""
从今天的预测报告文件导出Excel报告
只导出今天（2026-04-21）预测的比赛
"""
import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def find_today_prediction_files():
    """查找今天日期的预测报告文件"""
    today_str = datetime.now().strftime("%Y%m%d")
    pattern = f"prediction_*_{today_str}_*.txt"
    
    files = []
    for filename in os.listdir('.'):
        if filename.startswith('prediction_') and filename.endswith('.txt'):
            # 检查文件名中是否包含今天的日期
            if today_str in filename:
                files.append(filename)
    
    return sorted(files)

def parse_prediction_file(filepath):
    """解析单个预测报告文件，提取关键信息"""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    result = {
        "file": filepath,
        "competition": "未知",
        "time": "未知",
        "home_team": "未知",
        "away_team": "未知",
        "home_prob": 0,
        "draw_prob": 0,
        "away_prob": 0,
        "market": "未知",
        "handicap": "N/A",
        "recommendation": "未知",
        "value": 0,
        "confidence": 0,
        "risk_level": "未知",
        "upset_risk": 50
    }
    
    try:
        # 提取比赛信息（第6行）
        # 格式: "📅 比赛: 水晶宫 vs 西汉姆"
        match_line = None
        for line in content.split('\n'):
            if '比赛:' in line:
                match_line = line
                break
        
        if match_line:
            # 提取中文队名
            match_pattern = r'比赛:\s*(.+)\s+vs\s+(.+)'
            match_obj = re.search(match_pattern, match_line)
            if match_obj:
                result["home_team"] = match_obj.group(1).strip()
                result["away_team"] = match_obj.group(2).strip()
        
        # 提取联赛信息（第8行）
        # 格式: "📊 联赛: 英格兰超级联赛"
        league_line = None
        for line in content.split('\n'):
            if '联赛:' in line:
                league_line = line
                break
        
        if league_line:
            league_pattern = r'联赛:\s*(.+)'
            league_obj = re.search(league_pattern, league_line)
            if league_obj:
                league_name = league_obj.group(1).strip()
                result["competition"] = map_competition_name(league_name)
        
        # 提取隐含概率（第14-17行）
        # 格式: "  • 水晶宫 胜: 36.3%"
        prob_pattern = r'•\s*(.+?)\s+胜:\s*(\d+\.?\d*)%'
        prob_matches = re.findall(prob_pattern, content)
        
        for team, prob in prob_matches:
            prob_value = float(prob)
            if '主' in team or team == result["home_team"]:
                result["home_prob"] = prob_value
            elif '客' in team or team == result["away_team"]:
                result["away_prob"] = prob_value
            elif '平' in team:
                result["draw_prob"] = prob_value
        
        # 提取冷门风险（第19行）
        # 格式: "冷门风险: 高 (80/100)"
        risk_pattern = r'冷门风险:\s*(\S+)\s*\((\d+)/100\)'
        risk_match = re.search(risk_pattern, content)
        if risk_match:
            risk_level = risk_match.group(1)
            risk_score = int(risk_match.group(2))
            result["risk_level"] = risk_level
            result["upset_risk"] = risk_score
        
        # 提取最终推荐（第39-43行）
        # 格式: "推荐市场: 让球盘"
        #       "推荐选项: 谨慎或放弃"
        #       "💪 信心分数: 30%"
        market_pattern = r'推荐市场:\s*(.+)'
        market_match = re.search(market_pattern, content)
        if market_match:
            result["market"] = market_match.group(1).strip()
        
        recommendation_pattern = r'推荐选项:\s*(.+)'
        recommendation_match = re.search(recommendation_pattern, content)
        if recommendation_match:
            result["recommendation"] = recommendation_match.group(1).strip()
        
        confidence_pattern = r'信心分数:\s*(\d+\.?\d*)%'
        confidence_match = re.search(confidence_pattern, content)
        if confidence_match:
            result["confidence"] = float(confidence_match.group(1))
        
        # 设置时间（从文件名中提取或使用默认值）
        # 文件名格式: prediction_Crystal Palace_West Ham United_20260421_014122.txt
        filename = os.path.basename(filepath)
        time_match = re.search(r'_(\d{4})(\d{2})(\d{2})_(\d{2})(\d{2})(\d{2})\.', filename)
        if time_match:
            hour = time_match.group(4)
            minute = time_match.group(5)
            result["time"] = f"{hour}:{minute}"
        
        # 价值分简化：使用信心分数
        result["value"] = result["confidence"]
        
    except Exception as e:
        print(f"警告: 解析文件 {filepath} 时出错: {e}")
    
    return result

def map_competition_name(league_name):
    """将联赛中文名映射为简短的赛事名称"""
    mapping = {
        "英格兰超级联赛": "英超",
        "意大利甲级联赛": "意甲",
        "西班牙甲级联赛": "西甲",
        "德国甲级联赛": "德甲",
        "法国甲级联赛": "法甲",
        "法国乙级联赛": "法乙",
        "葡萄牙超级联赛": "葡超",
        "美国职业足球大联盟": "美职联",
        "日本职业足球联赛": "日职联",
        "欧洲冠军联赛": "欧冠",
        "欧罗巴联赛": "欧联杯",
        "欧协联": "欧会杯",
        "解放者杯": "解放者杯"
    }
    return mapping.get(league_name, league_name)

def export_to_excel(results, append_mode=True):
    """导出结果到Excel（支持追加模式）"""
    filename = f"足球预测_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(os.path.expanduser("~/Desktop"), filename)
    
    # 检查文件是否存在，如果存在则清空重新创建（根据用户要求只导出今天的）
    if os.path.exists(filepath):
        print(f"📂 发现现有文件，将创建新文件只包含今天的数据")
        append_mode = False
    
    if append_mode and os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
        # 找到下一个空行
        next_row = ws.max_row + 1
        print(f"📂 追加到现有文件，从第 {next_row} 行开始")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "足球预测"
        next_row = 4  # 数据从第4行开始（跳过标题和表头）
        
        # 设置标题行样式
        title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        
        # 设置表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        
        # 第1行：标题
        ws.merge_cells('A1:O1')
        title_cell = ws['A1']
        title_cell.value = f"⚽ 足球预测报告"
        title_cell.font = title_font
        title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        # 第2行：日期说明
        ws.merge_cells('A2:O2')
        date_cell = ws['A2']
        date_cell.value = f"最后更新: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        date_cell.font = Font(name='微软雅黑', size=10, italic=True)
        date_cell.alignment = Alignment(horizontal='right')
        
        # 第3行：表头
        headers = [
            "日期", "序号", "赛事", "时间", "主队", "客队",
            "主胜%", "平局%", "客胜%",
            "推荐市场", "盘口", "推荐选项",
            "价值分", "信心%", "风险等级", "推荐状态"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        ws.row_dimensions[3].height = 30
        
        # 设置列宽
        column_widths = {
            'A': 12,  # 日期
            'B': 6,   # 序号
            'C': 12,  # 赛事
            'D': 8,   # 时间
            'E': 16,  # 主队
            'F': 16,  # 客队
            'G': 8,   # 主胜%
            'H': 8,   # 平局%
            'I': 8,   # 客胜%
            'J': 10,  # 推荐市场
            'K': 10,  # 盘口
            'L': 16,  # 推荐选项
            'M': 8,   # 价值分
            'N': 8,   # 信心%
            'O': 10,  # 风险等级
            'P': 12,  # 推荐状态
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
    
    # 数据样式
    data_font = Font(name='微软雅黑', size=10)
    # 推荐场次 - 深绿色
    recommend_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    # 可关注 - 黄色
    watch_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    # 观望 - 灰色
    wait_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 按信心分数排序
    sorted_results = sorted(results, key=lambda x: x["confidence"], reverse=True)
    
    # 筛选推荐场次（最多3场，信心≥65%）
    recommended = [r for r in sorted_results if r["confidence"] >= 65][:3]
    
    # 如果高信心不足3场，取前2场（信心≥60%）
    if len(recommended) < 2:
        recommended = [r for r in sorted_results if r["confidence"] >= 60][:2]
    
    # 标记推荐状态
    for result in sorted_results:
        if result in recommended:
            result["status"] = "🎯 推荐"
        elif result["confidence"] >= 50:
            result["status"] = "👀 可关注"
        else:
            result["status"] = "⏸️ 观望"
    
    # 获取当前序号
    current_date = datetime.now().strftime('%Y-%m-%d')
    existing_count = sum(1 for row in ws.iter_rows(min_row=4, max_col=1, values_only=True) if row[0] == current_date)
    
    # 填充数据
    for i, result in enumerate(sorted_results, 1):
        row = next_row + i - 1
        
        # 日期
        ws.cell(row=row, column=1, value=current_date).alignment = center_align
        
        # 序号
        ws.cell(row=row, column=2, value=existing_count + i).alignment = center_align
        
        # 赛事
        ws.cell(row=row, column=3, value=result["competition"]).alignment = center_align
        
        # 时间
        ws.cell(row=row, column=4, value=result["time"]).alignment = center_align
        
        # 主队
        ws.cell(row=row, column=5, value=result["home_team"]).alignment = center_align
        
        # 客队
        ws.cell(row=row, column=6, value=result["away_team"]).alignment = center_align
        
        # 隐含概率
        ws.cell(row=row, column=7, value=result["home_prob"]).alignment = center_align
        ws.cell(row=row, column=8, value=result["draw_prob"]).alignment = center_align
        ws.cell(row=row, column=9, value=result["away_prob"]).alignment = center_align
        
        # 推荐市场
        ws.cell(row=row, column=10, value=result["market"]).alignment = center_align
        
        # 盘口（预测报告中没有盘口数据，留空）
        ws.cell(row=row, column=11, value=result["handicap"]).alignment = center_align
        
        # 推荐选项
        ws.cell(row=row, column=12, value=result["recommendation"]).alignment = center_align
        
        # 价值分
        ws.cell(row=row, column=13, value=result["value"]).alignment = center_align
        
        # 信心分数
        conf_cell = ws.cell(row=row, column=14, value=result["confidence"])
        conf_cell.alignment = center_align
        
        # 风险等级
        risk_cell = ws.cell(row=row, column=15, value=result["risk_level"])
        risk_cell.alignment = center_align
        
        # 推荐状态
        status_cell = ws.cell(row=row, column=16, value=result["status"])
        status_cell.alignment = center_align
        
        # 根据推荐状态设置背景色
        if result["status"] == "🎯 推荐":
            row_fill = recommend_fill
        elif result["status"] == "👀 可关注":
            row_fill = watch_fill
        else:
            row_fill = wait_fill
        
        # 应用样式到整行
        for col in range(1, 17):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border
        
        ws.row_dimensions[row].height = 25
    
    # 更新最后更新时间
    if ws.max_row >= 2:
        ws.cell(row=2, column=1).value = f"最后更新: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # 保存文件
    wb.save(filepath)
    
    print(f"✅ Excel报告已保存: {filepath}")
    return filepath, sorted_results

def main():
    print("=" * 60)
    print("📊 从今天预测报告导出Excel（仅今天数据）")
    print("=" * 60)
    
    # 查找今天的预测文件
    today_files = find_today_prediction_files()
    
    if not today_files:
        print("❌ 未找到今天日期的预测报告文件")
        return
    
    print(f"✅ 找到 {len(today_files)} 个今天预测的文件:")
    for f in today_files:
        print(f"   • {f}")
    
    # 解析所有文件
    results = []
    for filepath in today_files:
        print(f"\n解析文件: {filepath}")
        result = parse_prediction_file(filepath)
        results.append(result)
        
        print(f"   {result['home_team']} vs {result['away_team']}")
        print(f"   联赛: {result['competition']}, 信心: {result['confidence']}%")
        print(f"   推荐: {result['recommendation']}")
    
    # 导出Excel
    filepath, sorted_results = export_to_excel(results, append_mode=False)
    
    # 打印汇总
    print("\n" + "=" * 60)
    print("📊 今天预测汇总")
    print("=" * 60)
    
    # 筛选推荐场次
    recommended = [r for r in sorted_results if r["confidence"] >= 65][:3]
    if len(recommended) < 2:
        recommended = [r for r in sorted_results if r["confidence"] >= 60][:2]
    
    watch = [r for r in sorted_results if r not in recommended and r["confidence"] >= 50]
    wait = [r for r in sorted_results if r["confidence"] < 50]
    
    print(f"\n🎯 今日推荐 ({len(recommended)}场):")
    for r in recommended:
        print(f"   ⭐ {r['home_team']} vs {r['away_team']}")
        print(f"      推荐: {r['recommendation']} | 信心: {r['confidence']}%")
    
    if watch:
        print(f"\n👀 可关注 ({len(watch)}场):")
        for r in watch:
            print(f"   • {r['home_team']} vs {r['away_team']}: {r['recommendation']} ({r['confidence']}%)")
    
    if wait:
        print(f"\n⏸️ 观望 ({len(wait)}场):")
        for r in wait:
            print(f"   • {r['home_team']} vs {r['away_team']}: {r['confidence']}%")
    
    print(f"\n📁 Excel文件已保存: {filepath}")
    print(f"📊 共 {len(sorted_results)} 场比赛, 推荐 {len(recommended)} 场")

if __name__ == "__main__":
    main()