#!/usr/bin/env python3
"""
批量预测脚本 - 预测多场比赛并导出Excel
"""
from football_predictor import FootballPredictor
from team_translator import translate_team_name
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os

# 所有比赛列表
MATCHES = [
    # 欧联杯
    {"home": "Nottingham Forest", "away": "Porto", "league": "soccer_uefa_europa_league", "time": "03:00", "competition": "欧联杯"},
    {"home": "Real Betis", "away": "SC Braga", "league": "soccer_uefa_europa_league", "time": "03:00", "competition": "欧联杯"},
    {"home": "Aston Villa", "away": "Bologna", "league": "soccer_uefa_europa_league", "time": "03:00", "competition": "欧联杯"},

    # 欧会杯
    {"home": "Fiorentina", "away": "Crystal Palace", "league": "soccer_uefa_europa_conference_league", "time": "03:00", "competition": "欧会杯"},
    {"home": "Strasbourg", "away": "FSV Mainz 05", "league": "soccer_uefa_europa_conference_league", "time": "03:00", "competition": "欧会杯"},
    {"home": "AEK Athens", "away": "Rayo Vallecano", "league": "soccer_uefa_europa_conference_league", "time": "03:00", "competition": "欧会杯"},

    # 解放者杯
    {"home": "Palmeiras-SP", "away": "Sporting Cristal", "league": "soccer_conmebol_copa_libertadores", "time": "06:00", "competition": "解放者杯"},
    {"home": "Flamengo-RJ", "away": "Independiente Medellín", "league": "soccer_conmebol_copa_libertadores", "time": "08:30", "competition": "解放者杯"},
]

def run_predictions():
    """运行所有预测并收集结果"""
    predictor = FootballPredictor()
    results = []

    print("=" * 60)
    print("🏆 批量比赛预测")
    print("=" * 60)

    for i, match in enumerate(MATCHES, 1):
        home_cn = translate_team_name(match["home"])
        away_cn = translate_team_name(match["away"])

        print(f"\n[{i}/{len(MATCHES)}] 预测: {home_cn} vs {away_cn}")

        try:
            result = predictor.predict(match["home"], match["away"], match["league"])

            if result:
                # 提取关键信息
                consensus = predictor.reports
                asian_value = consensus.get("asian", {}).get("value_score", 0)
                overunder_value = consensus.get("overunder", {}).get("overunder_value_score", 0)
                upset_risk = consensus.get("upset", {}).get("upset_risk_score", 0)

                # 获取隐含概率
                stats_prob = consensus.get("stats", {}).get("implied_probability", {})
                home_prob = stats_prob.get("home", 0)
                draw_prob = stats_prob.get("draw", 0)
                away_prob = stats_prob.get("away", 0)

                # 获取最终推荐
                if asian_value > overunder_value:
                    market = "让球盘"
                    recommendation = consensus.get("asian", {}).get("recommendation", "N/A")
                    value = asian_value
                    handicap = consensus.get("asian", {}).get("actual_handicap", "N/A")
                else:
                    market = "大小球"
                    bias = consensus.get("overunder", {}).get("market_bias", "N/A")
                    line = consensus.get("overunder", {}).get("mainstream_total_line", 2.5)
                    recommendation = f"{bias} {line}球"
                    value = overunder_value
                    handicap = f"{line}球"

                # 获取信心分数
                base_score = (
                    consensus.get("stats", {}).get("confidence_weight", 50) * 0.4 +
                    consensus.get("tactics", {}).get("tactical_match_score", 50) * 0.25 +
                    consensus.get("sentiment", {}).get("market_sentiment_score", 50) * 0.2 +
                    (100 - upset_risk) * 0.15
                )
                market_factor = value / 100 if value > 0 else 0.5
                final_confidence = base_score * (0.7 + 0.3 * market_factor)
                final_confidence = max(30, min(85, final_confidence))

                # 风险等级
                if upset_risk >= 60:
                    risk_level = "高"
                elif upset_risk >= 40:
                    risk_level = "中"
                else:
                    risk_level = "低"

                results.append({
                    "competition": match["competition"],
                    "time": match["time"],
                    "home_team": home_cn,
                    "away_team": away_cn,
                    "home_prob": round(home_prob, 1),
                    "draw_prob": round(draw_prob, 1),
                    "away_prob": round(away_prob, 1),
                    "market": market,
                    "handicap": handicap,
                    "recommendation": recommendation,
                    "value": value,
                    "confidence": round(final_confidence, 1),
                    "risk_level": risk_level,
                    "upset_risk": upset_risk
                })
                print(f"   ✅ 推荐: {recommendation} | 信心: {final_confidence:.1f}%")
            else:
                print(f"   ❌ 无数据")
                results.append({
                    "competition": match["competition"],
                    "time": match["time"],
                    "home_team": home_cn,
                    "away_team": away_cn,
                    "home_prob": 0,
                    "draw_prob": 0,
                    "away_prob": 0,
                    "market": "N/A",
                    "handicap": "N/A",
                    "recommendation": "无数据",
                    "value": 0,
                    "confidence": 0,
                    "risk_level": "N/A",
                    "upset_risk": 0
                })

        except Exception as e:
            print(f"   ❌ 错误: {e}")
            results.append({
                "competition": match["competition"],
                "time": match["time"],
                "home_team": home_cn,
                "away_team": away_cn,
                "home_prob": 0,
                "draw_prob": 0,
                "away_prob": 0,
                "market": "N/A",
                "handicap": "N/A",
                "recommendation": "预测失败",
                "value": 0,
                "confidence": 0,
                "risk_level": "N/A",
                "upset_risk": 0
            })

    return results

def export_to_excel(results, append_mode=True):
    """导出结果到Excel（支持追加模式）"""
    from openpyxl import load_workbook

    filename = f"足球预测_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(os.path.expanduser("~/Desktop"), filename)

    # 检查文件是否存在，如果存在则追加
    if append_mode and os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
        # 找到下一个空行
        next_row = ws.max_row + 1
        print(f"📂 追加到现有文件，从第 {next_row} 行开始")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "足球预测"
        next_row = 4  # 数据从第4行开始（跳过标题和表头）

        # 设置标题行样式
        title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

        # 设置表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')

        # 第1行：标题
        ws.merge_cells('A1:O1')
        title_cell = ws['A1']
        title_cell.value = f"⚽ 足球预测报告"
        title_cell.font = title_font
        title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40

        # 第2行：日期说明
        ws.merge_cells('A2:O2')
        date_cell = ws['A2']
        date_cell.value = f"最后更新: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        date_cell.font = Font(name='微软雅黑', size=10, italic=True)
        date_cell.alignment = Alignment(horizontal='right')

        # 第3行：表头
        headers = [
            "日期", "序号", "赛事", "时间", "主队", "客队",
            "主胜%", "平局%", "客胜%",
            "推荐市场", "盘口", "推荐选项",
            "价值分", "信心%", "风险等级", "推荐状态"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        ws.row_dimensions[3].height = 30

        # 设置列宽
        column_widths = {
            'A': 12,  # 日期
            'B': 6,   # 序号
            'C': 12,  # 赛事
            'D': 8,   # 时间
            'E': 16,  # 主队
            'F': 16,  # 客队
            'G': 8,   # 主胜%
            'H': 8,   # 平局%
            'I': 8,   # 客胜%
            'J': 10,  # 推荐市场
            'K': 10,  # 盘口
            'L': 16,  # 推荐选项
            'M': 8,   # 价值分
            'N': 8,   # 信心%
            'O': 10,  # 风险等级
            'P': 12,  # 推荐状态
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    # 数据样式
    data_font = Font(name='微软雅黑', size=10)
    # 推荐场次 - 深绿色
    recommend_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    # 可关注 - 黄色
    watch_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    # 观望 - 灰色
    wait_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # 按信心分数排序
    sorted_results = sorted(results, key=lambda x: x["confidence"], reverse=True)

    # 筛选推荐场次（最多3场，信心≥65%）
    recommended = [r for r in sorted_results if r["confidence"] >= 65][:3]

    # 如果高信心不足3场，取前2场（信心≥60%）
    if len(recommended) < 2:
        recommended = [r for r in sorted_results if r["confidence"] >= 60][:2]

    # 标记推荐状态
    for result in sorted_results:
        if result in recommended:
            result["status"] = "🎯 推荐"
        elif result["confidence"] >= 50:
            result["status"] = "👀 可关注"
        else:
            result["status"] = "⏸️ 观望"

    # 获取当前序号
    current_date = datetime.now().strftime('%Y-%m-%d')
    existing_count = sum(1 for row in ws.iter_rows(min_row=4, max_col=1, values_only=True) if row[0] == current_date)

    # 填充数据
    for i, result in enumerate(sorted_results, 1):
        row = next_row + i - 1

        # 日期
        ws.cell(row=row, column=1, value=current_date).alignment = center_align

        # 序号
        ws.cell(row=row, column=2, value=existing_count + i).alignment = center_align

        # 赛事
        ws.cell(row=row, column=3, value=result["competition"]).alignment = center_align

        # 时间
        ws.cell(row=row, column=4, value=result["time"]).alignment = center_align

        # 主队
        ws.cell(row=row, column=5, value=result["home_team"]).alignment = center_align

        # 客队
        ws.cell(row=row, column=6, value=result["away_team"]).alignment = center_align

        # 隐含概率
        ws.cell(row=row, column=7, value=result["home_prob"]).alignment = center_align
        ws.cell(row=row, column=8, value=result["draw_prob"]).alignment = center_align
        ws.cell(row=row, column=9, value=result["away_prob"]).alignment = center_align

        # 推荐市场
        ws.cell(row=row, column=10, value=result["market"]).alignment = center_align

        # 盘口
        ws.cell(row=row, column=11, value=result["handicap"]).alignment = center_align

        # 推荐选项
        ws.cell(row=row, column=12, value=result["recommendation"]).alignment = center_align

        # 价值分
        ws.cell(row=row, column=13, value=result["value"]).alignment = center_align

        # 信心分数
        conf_cell = ws.cell(row=row, column=14, value=result["confidence"])
        conf_cell.alignment = center_align

        # 风险等级
        risk_cell = ws.cell(row=row, column=15, value=result["risk_level"])
        risk_cell.alignment = center_align

        # 推荐状态
        status_cell = ws.cell(row=row, column=16, value=result["status"])
        status_cell.alignment = center_align

        # 根据推荐状态设置背景色
        if result["status"] == "🎯 推荐":
            row_fill = recommend_fill
        elif result["status"] == "👀 可关注":
            row_fill = watch_fill
        else:
            row_fill = wait_fill

        # 应用样式到整行
        for col in range(1, 17):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border

        ws.row_dimensions[row].height = 25

    # 更新最后更新时间
    if ws.max_row >= 2:
        ws.cell(row=2, column=1).value = f"最后更新: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # 保存文件
    wb.save(filepath)

    print(f"\n✅ Excel报告已保存: {filepath}")
    return filepath

def main():
    # 运行预测
    results = run_predictions()

    # 导出Excel
    filepath = export_to_excel(results)

    # 打印汇总
    print("\n" + "=" * 60)
    print("📊 预测汇总")
    print("=" * 60)

    sorted_results = sorted(results, key=lambda x: x["confidence"], reverse=True)

    # 筛选推荐场次（最多3场，信心≥65%）
    recommended = [r for r in sorted_results if r["confidence"] >= 65][:3]
    if len(recommended) < 2:
        recommended = [r for r in sorted_results if r["confidence"] >= 60][:2]

    watch = [r for r in sorted_results if r not in recommended and r["confidence"] >= 50]
    wait = [r for r in sorted_results if r["confidence"] < 50]

    print(f"\n🎯 今日推荐 ({len(recommended)}场):")
    for r in recommended:
        print(f"   ⭐ {r['home_team']} vs {r['away_team']}")
        print(f"      推荐: {r['recommendation']} | 信心: {r['confidence']}%")

    if watch:
        print(f"\n👀 可关注 ({len(watch)}场):")
        for r in watch:
            print(f"   • {r['home_team']} vs {r['away_team']}: {r['recommendation']} ({r['confidence']}%)")

    if wait:
        print(f"\n⏸️ 观望 ({len(wait)}场):")
        for r in wait:
            print(f"   • {r['home_team']} vs {r['away_team']}: {r['confidence']}%")

    print(f"\n📁 Excel文件已保存: {filepath}")
    print(f"📊 共 {len(sorted_results)} 场比赛, 推荐 {len(recommended)} 场")

if __name__ == "__main__":
    main()
